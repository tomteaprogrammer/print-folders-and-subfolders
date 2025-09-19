#!/usr/bin/env python3
"""
Usage:
    python export_folders_progress_ui_winfix.py

Improvements for Windows:
    - Option to use the extended-length path prefix (\\?\\) to reduce MAX_PATH issues.
    - Explicit handling and logging for WinError 3 ("The system cannot find the path specified").
    - Still skips files and does not download online-only contents.
"""

import os
import sys
import csv
import time
import errno
import subprocess
from pathlib import Path
from typing import List, Tuple

# GUI imports
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception as e:
    print("tkinter is required for the GUI. Install it or run on a system with tkinter available.", file=sys.stderr)
    raise

IS_WINDOWS = (os.name == "nt")

# Track whether we already attempted an install to avoid loops
_ALREADY_TRIED_INSTALL = os.environ.get("OPENPYXL_INSTALL_ATTEMPTED") == "1"

# Try to import openpyxl; if missing, attempt install and self-restart exactly once
HAVE_OPENPYXL = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:
    if not _ALREADY_TRIED_INSTALL:
        try:
            print("openpyxl not found. Installing...")
            subprocess.check_check_call  # force NameError if older python
        except Exception:
            pass
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            env = dict(os.environ)
            env["OPENPYXL_INSTALL_ATTEMPTED"] = "1"
            print("openpyxl installed. Restarting script...")
            os.execve(sys.executable, [sys.executable] + sys.argv, env)
        except Exception as e:
            print("Failed to install openpyxl:", e)
            HAVE_OPENPYXL = False
    else:
        HAVE_OPENPYXL = False


def write_excel(rows: List[List[str]], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    max_levels = max(len(r) for r in rows) if rows else 1
    cols = [f"Level{i}" for i in range(1, max_levels + 1)]
    norm_rows = [r + [""] * (max_levels - len(r)) for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "Folders"

    header_font = Font(bold=True)
    fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = left
        cell.border = border

    for row in norm_rows:
        ws.append(row)

    for r in range(2, 2 + len(norm_rows)):
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = left
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{1 + len(norm_rows)}"

    for c in range(1, len(cols) + 1):
        col_letter = get_column_letter(c)
        max_len = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, len(norm_rows) + 2))
        ws.column_dimensions[col_letter].width = max(12, min(60, max_len + 2))

    wb.save(out_path.as_posix())


def write_csv(rows: List[List[str]], out_path: Path) -> None:
    max_levels = max(len(r) for r in rows) if rows else 1
    cols = [f"Level{i}" for i in range(1, max_levels + 1)]
    norm_rows = [r + [""] * (max_levels - len(r)) for r in rows]

    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(cols)
        writer.writerows(norm_rows)


def save_log(skipped: List[Tuple[str, str]], log_path: Path) -> None:
    with log_path.open("w", encoding="utf-8") as f:
        f.write("Skipped folders log\n")
        f.write("===================\n")
        if not skipped:
            f.write("No skipped folders.\n")
            return
        for path, reason in skipped:
            f.write(f"{path} | {reason}\n")


def to_long_path(p: Path, enable: bool) -> Path:
    if IS_WINDOWS and enable:
        s = str(p)
        # If already long-path, return as-is
        if s.startswith('\\\\?\\'):
            return p
        # UNC paths need \\?\UNC\ prefix
        if s.startswith('\\\\'):
            long_s = '\\\\?\\UNC' + s[1:]
        else:
            long_s = '\\\\?\\' + s
        return Path(long_s)
    return p
def scan_folders_with_progress(root: Path, use_long_paths: bool, progress_win, lbl_count, lbl_path, cancel_state) -> Tuple[List[List[str]], List[Tuple[str, str]]]:
    """Scan folders recursively while updating a small progress window. Supports Cancel."""
    rows: List[List[str]] = []
    skipped: List[Tuple[str, str]] = []

    root_lp = to_long_path(root, use_long_paths)
    rows.append([root.name])
    found = 1
    last_update = 0.0
    lbl_count.config(text=f"Folders found: {found}")
    lbl_path.config(text=f"Current: {root.name}")
    progress_win.update_idletasks()

    def walk_dir(dir_path: Path, ancestors: List[str]):
        nonlocal found, last_update
        if cancel_state["cancel"]:
            return
        try:
            # Use long path where applicable
            scan_target = to_long_path(dir_path, use_long_paths)
            with os.scandir(str(scan_target)) as it:
                for entry in it:
                    if cancel_state["cancel"]:
                        return
                    try:
                        if not entry.is_dir(follow_symlinks=False):
                            continue
                    except OSError as e:
                        # If WinError 3, log explicitly
                        if hasattr(e, "winerror") and e.winerror == 3:
                            skipped.append((entry.path, "WinError 3 (path not found) - likely placeholder/online-only or moved"))
                        else:
                            skipped.append((entry.path, f"is_dir failed: {e}"))
                        continue

                    name = entry.name
                    row = ancestors + [name]
                    rows.append(row)
                    found += 1

                    now = time.time()
                    if now - last_update > 0.05:
                        lbl_count.config(text=f"Folders found: {found}")
                        lbl_path.config(text=f"Current: {'/'.join(row[-4:])}")
                        progress_win.update_idletasks()
                        last_update = now

                    try:
                        walk_dir(Path(entry.path), row)
                    except RecursionError as e:
                        skipped.append((entry.path, f"recursion error: {e}"))
        except OSError as e:
            # Catch directory-level failures
            if hasattr(e, "winerror") and e.winerror == 3:
                skipped.append((str(dir_path), "WinError 3 (path not found) - likely placeholder/online-only or moved"))
            elif e.errno == errno.ENOENT:
                skipped.append((str(dir_path), "ENOENT (no such file or directory)"))
            elif e.errno == errno.EACCES:
                skipped.append((str(dir_path), "EACCES (permission denied)"))
            else:
                skipped.append((str(dir_path), f"os error: {e}"))

    walk_dir(root_lp, [root.name])
    return rows, skipped


def main():
    root = tk.Tk()
    root.withdraw()

    messagebox.showinfo("Pick Folder", "Select your Dropbox/OneDrive/local folder. The script lists folder names only.")

    folder = filedialog.askdirectory(title="Choose folder")
    if not folder:
        messagebox.showwarning("Canceled", "No folder selected.")
        return

    base = Path(folder).resolve()
    if not base.exists() or not base.is_dir():
        messagebox.showerror("Error", "The selected path is not a folder.")
        return

    # Ask for save location BEFORE scanning
    default_xlsx = base.name + "_folders.xlsx"
    default_csv = base.name + "_folders.csv"
    want_xlsx = HAVE_OPENPYXL

    out_name = filedialog.asksaveasfilename(
        title="Save output as",
        initialdir=base.parent.as_posix(),
        initialfile=default_xlsx if want_xlsx else default_csv,
        defaultextension=".xlsx" if want_xlsx else ".csv",
        filetypes=[("Excel Workbook", "*.xlsx"), ("CSV (Comma delimited)", "*.csv"), ("All files", "*.*")],
    )
    if not out_name:
        messagebox.showwarning("Canceled", "No output file chosen.")
        return

    out_path = Path(out_name)

    # Small options dialog for Windows long paths
    use_long_paths = True if IS_WINDOWS else False
    if IS_WINDOWS:
        opt = tk.Toplevel()
        opt.title("Options")
        var_long = tk.BooleanVar(value=True)
        tk.Checkbutton(opt, text="Use Windows long-path prefix (\\\\?\\)", variable=var_long).pack(anchor="w", padx=10, pady=10)
        def go():
            nonlocal use_long_paths
            use_long_paths = var_long.get()
            opt.destroy()
        tk.Button(opt, text="Start Scan", command=go).pack(pady=(0,10))
        opt.grab_set()
        opt.wait_window()

    # Build progress window
    progress = tk.Toplevel()
    progress.title("Scanning...")
    progress.geometry("480x140")
    lbl_count = tk.Label(progress, text="Folders found: 0", anchor="w", justify="left")
    lbl_count.pack(fill="x", padx=10, pady=(10, 2))
    lbl_path = tk.Label(progress, text="Current:", anchor="w", justify="left")
    lbl_path.pack(fill="x", padx=10, pady=2)

    cancel_state = {"cancel": False}
    def do_cancel():
        cancel_state["cancel"] = True
        lbl_path.config(text="Cancelling...")

    btn_cancel = tk.Button(progress, text="Cancel", command=do_cancel)
    btn_cancel.pack(pady=10)

    progress.update_idletasks()

    # Scan with UI
    rows, skipped = scan_folders_with_progress(base, use_long_paths, progress, lbl_count, lbl_path, cancel_state)

    # Close progress window
    try:
        progress.destroy()
    except Exception:
        pass

    if cancel_state["cancel"] and len(rows) <= 1:
        messagebox.showinfo("Canceled", "Scan canceled. No data saved.")
        return

    # Save data
    try:
        if out_path.suffix.lower() == ".xlsx" and HAVE_OPENPYXL:
            write_excel(rows, out_path)
            saved_path = out_path
        else:
            csv_path = out_path if out_path.suffix.lower() == ".csv" else out_path.with_suffix(".csv")
            write_csv(rows, csv_path)
            saved_path = csv_path
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save output:\n{e}")
        raise

    # Save log
    log_path = saved_path.with_suffix(".scan_log.txt")
    try:
        save_log(skipped, log_path)
    except Exception as e:
        messagebox.showwarning("Log", f"Could not write log file:\n{e}")
        log_path = None

    # Summary
    found_count = len(rows)
    skipped_count = len(skipped)
    msg = f"Saved {saved_path}\n\nFolders found: {found_count}\nSkipped: {skipped_count}"
    if log_path:
        msg += f"\nLog: {log_path}"
    if skipped_count > 0:
        msg += "\nCommon reasons: online-only placeholders, moved/renamed during scan, or long-path/permission limits."
    if cancel_state["cancel"]:
        msg += "\nNote: Scan was canceled early; results are partial."
    messagebox.showinfo("Done", msg)


if __name__ == "__main__":
    main()
